# -*- coding: utf-8 -*-
# ==============================================================================
#  insurance.mouvement.report.wizard — Rapport Mouvement Clients Primes
# ==============================================================================

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import re
from pathlib import Path
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class InsuranceMouvementReportWizard(models.TransientModel):
    """
    Assistant pour générer l'état 'Mouvement Clients primes' sous format Excel (.xlsx).
    Extrait les données en temps réel depuis le dump de base legacy (DATA_ASS.txt)
    avec un tri strict identique à l'ancienne requête Oracle (DATE_OP, NUM_OPERATION).
    """
    _name = 'insurance.mouvement.report.wizard'
    _description = 'Assistant Rapport Mouvement Clients Primes'

    excel_file = fields.Binary(string='Fichier Excel', readonly=True)
    excel_filename = fields.Char(string='Nom du fichier', default='Etat Mouvement Clients primes.xlsx')
    state = fields.Selection([
        ('draft', 'Prêt à générer'),
        ('done', 'Prêt pour le téléchargement')
    ], default='draft', readonly=True)

    def _get_data_file_path(self):
        """Recherche le fichier DATA_ASS.txt de manière sécurisée et robuste."""
        import odoo.modules as odoo_modules
        path_str = odoo_modules.get_module_resource('assurcore', 'data', 'DATA_ASS.txt')
        if path_str and Path(path_str).exists():
            return Path(path_str)
            
        fallback_paths = [
            Path('/mnt/extra-addons/assurcore/data/DATA_ASS.txt'),
            Path('/Robot/ASSURPROD/DATA_ASS.txt'),
            Path('d:/Robot/ASSURPROD/DATA_ASS.txt'),
            Path('D:/Robot/ASSURPROD/DATA_ASS.txt'),
        ]
        for p in fallback_paths:
            if p.exists():
                return p
        raise UserError(_("Le fichier de données DATA_ASS.txt est introuvable dans le dossier 'data' du module."))

    def action_generate_excel(self):
        """Lit, filtre, ordonne les données et génère le fichier Excel premium."""
        filepath = self._get_data_file_path()
        
        SEPARATOR = '\t'
        ENCODING = 'utf-8'
        
        rows = []
        headers = []
        
        try:
            with open(filepath, 'r', encoding=ENCODING, errors='replace') as f:
                in_block = False
                for line in f:
                    line = line.rstrip('\n\r')
                    parts = [p.strip().strip('"') for p in line.split(SEPARATOR)]
                    if not any(parts):
                        continue
                    
                    if not in_block:
                        if 'NUM_OPERATION' in parts and 'NUM_POLICE' in parts and 'ANNEE_FACT_PRIME' in parts:
                            in_block = True
                            headers = parts
                            continue
                    else:
                        upper_cnt = sum(
                            1 for p in parts
                            if p == p.upper() and re.match(r'^[A-Z][A-Z0-9_]*$', p) and len(p) > 1
                        )
                        if upper_cnt >= len(parts) * 0.6 and len(parts) >= 2 and 'NUM_OPERATION' not in parts:
                            break
                        
                        if len(parts) != len(headers):
                            continue
                        
                        row = dict(zip(headers, parts))
                        rows.append(row)
        except Exception as exc:
            raise UserError(_("Erreur lors de la lecture du fichier : %s") % exc)
            
        if not rows:
            raise UserError(_("Aucune donnée trouvée dans le fichier de données pour le bloc PR_OPERATION."))

        # Fonctions de conversion pour le tri et l'affichage
        def parse_date_obj(value):
            if not value or value.strip() in ('', 'NULL', 'null'):
                return date(1970, 1, 1)
            v = value.strip()
            # DD/MM/YY
            m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', v)
            if m:
                day, month, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                year = 2000 + yy if yy < 50 else 1900 + yy
                try: return date(year, month, day)
                except ValueError: pass
            # DD/MM/YYYY
            m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', v)
            if m:
                day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
                try: return date(year, month, day)
                except ValueError: pass
            # YYYY-MM-DD
            m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', v)
            if m:
                try: return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                except ValueError: pass
            return date(1970, 1, 1)

        def parse_int_val(value):
            try:
                v = value.strip().replace(' ', '')
                return int(v)
            except:
                return 0

        # Tri identique Oracle: ORDER BY DATE_OP (col 3), NUM_OPERATION (col 2)
        rows.sort(key=lambda r: (parse_date_obj(r.get('DATE_OP', '')), parse_int_val(r.get('NUM_OPERATION', ''))))

        # Génération du classeur Excel openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mvt Clients Primes"
        ws.views.sheetView[0].showGridLines = True

        # Styles premium (Thème Navy & Steel Blue)
        title_font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
        meta_font = Font(name='Calibri', size=10, italic=True, color='5D6D7E')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Calibri', size=10)
        total_font = Font(name='Calibri', size=11, bold=True, color='1F4E78')
        
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')
        total_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='double', color='1F4E78')
        )

        # En-tête de page
        ws.merge_cells('A1:W1')
        ws['A1'] = "Etat Mouvement Clients primes"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35
        
        ws['A2'] = f"Exporté le : {date.today().strftime('%d/%m/%Y')} | Source : Oracle Legacy (DATA_ASS) | Total : {len(rows)} opérations"
        ws['A2'].font = meta_font
        ws.row_dimensions[2].height = 18
        
        ws.row_dimensions[3].height = 8

        # Titres des colonnes requis
        headers_labels = [
            "N° Police", "N° Opération", "Date Opération", "Désignation",
            "N° Quittance", "N° Attestation", "Véhicule", "Compagnie",
            "Du", "Au", "Montant Prime", "Honoraires HT",
            "Année Fact. Prime", "Réf Facture Prime", "Année Fact. Hon.", "Réf Facture Hon.",
            "Code Client", "Cat. Facture Prime", "Cat. Facture Hon.", "Attribut Client",
            "Type Client", "Raison Sociale", "T_C"
        ]
        
        keys = [
            'NUM_POLICE', 'NUM_OPERATION', 'DATE_OP', 'DESIGNATION',
            'NUM_QUITTANCE', 'NUM_ATTESTATION', 'VEHICULE', 'COMPAGNIE',
            'DATE_VALIDITE_DU', 'DATE_VALIDITE_AU', 'MONTANT_PRIME', 'MONTANT_HONORAIRE_HT',
            'ANNEE_FACT_PRIME', 'NUM_EDIT_FACTURE_PRIME', 'ANNEE_FACT_HON', 'NUM_EDIT_FACTURE_HON',
            'NUM_CLIENT', 'CATEGORIE_FACTURE_PRIME', 'CATEGORIE_FACTURE_HON', 'ATTRIBUT_CLIENT',
            'TYPE_CLIENT', 'RAISON_SOCIALE', 'T_C'
        ]

        for col_idx, label in enumerate(headers_labels, 1):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[4].height = 28

        # Remplissage des lignes de données
        current_row = 5
        for row_idx, r in enumerate(rows):
            for col_idx, key in enumerate(keys, 1):
                val = r.get(key, '')
                if key in ('MONTANT_PRIME', 'MONTANT_HONORAIRE_HT'):
                    try:
                        val = float(val.replace(',', '.').replace(' ', ''))
                    except:
                        val = 0.0
                elif key in ('ANNEE_FACT_PRIME', 'ANNEE_FACT_HON'):
                    try:
                        val = int(val)
                    except:
                        pass
                
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                
                # Formatage et alignements spécifiques
                if key in ('MONTANT_PRIME', 'MONTANT_HONORAIRE_HT'):
                    cell.number_format = '#,##0.000'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif key in ('DATE_OP', 'DATE_VALIDITE_DU', 'DATE_VALIDITE_AU'):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif key in ('NUM_OPERATION', 'ANNEE_FACT_PRIME', 'ANNEE_FACT_HON', 'NUM_CLIENT', 'T_C', 'TYPE_CLIENT'):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Ligne de Totalisation
        total_row = current_row
        ws.cell(row=total_row, column=1, value="Total Général").font = total_font
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=total_row, column=1).border = double_bottom_border
        
        for col_idx in range(2, len(keys) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.border = double_bottom_border
            
        prime_col_letter = openpyxl.utils.get_column_letter(11)
        hon_col_letter = openpyxl.utils.get_column_letter(12)
        
        prime_total_cell = ws.cell(row=total_row, column=11, value=f"=SUM({prime_col_letter}5:{prime_col_letter}{total_row-1})")
        prime_total_cell.font = total_font
        prime_total_cell.number_format = '#,##0.000'
        prime_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        prime_total_cell.fill = total_fill
        
        hon_total_cell = ws.cell(row=total_row, column=12, value=f"=SUM({hon_col_letter}5:{hon_col_letter}{total_row-1})")
        hon_total_cell.font = total_font
        hon_total_cell.number_format = '#,##0.000'
        hon_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        hon_total_cell.fill = total_fill
        
        ws.row_dimensions[total_row].height = 25

        # Filtre automatique
        ws.auto_filter.ref = f"A4:W{total_row-1}"

        # Ajustement des largeurs de colonnes
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Enregistrement en Bytes et conversion base64
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        excel_bytes = fp.read()
        fp.close()

        self.write({
            'excel_file': base64.b64encode(excel_bytes),
            'excel_filename': 'Etat Mouvement Clients primes.xlsx',
            'state': 'done'
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_draft(self):
        """Réinitialise l'assistant pour une nouvelle génération."""
        self.write({'state': 'draft', 'excel_file': False})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
